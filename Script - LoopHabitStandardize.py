# HabitLoopStandardize.py
# Programme processes data into a standardized format which enables it to be easily analysed at a later date.

# import relevant modules
import pyinputplus as pyip
from pathlib import Path
import csv
import openpyxl

print('Programme started')

# input file path of csv file to be processed
windowsFP = pyip.inputFilepath(prompt='Input file path of csv file to be processed: \n', strip='"',
                               blank=False, mustExist=True)

# get csv data file path
p = Path(windowsFP)
parent = p.parent

# create file path for xlsx version of original data - xlsx
stem = p.stem + ' - Original.xlsx'
xlsxfilepath = Path(parent, stem)

# create file path for processed data - xlsx 2
editedstem = p.stem + ' - Edited.xlsx'
editedfilepath = Path(parent, editedstem)

print('Relevant file paths created')

# User inputs the earliest date to which they want the information processed to
lowerdate = pyip.inputRegex(r'\d{4}-\d{2}-\d{2}', prompt="Input lower bound date yyyy-mm-dd for "
                                                         "programme to process information to. "
                                                         "Enter blank to process all information: \n", blank=True)

print('Data is being processed')

# convert and save csv to xlsx file
wb = openpyxl.Workbook()
ws = wb.active

with open(p) as f:
    reader = csv.reader(f, delimiter=',', )
    for row in reader:
        ws.append(row)
wb.save(xlsxfilepath)

# Load the original data xlsx file
xlsxwb = openpyxl.load_workbook(xlsxfilepath)
xlsxws = xlsxwb.active

# create an additional xlsx file to input the processed information into - xlsx 2
xlsx2 = openpyxl.Workbook()
xlsxws2 = xlsx2.active

# cycle through rows in original xlsx file and input processed information into the xlsx 2 file
for rownum in range(xlsxws.max_row):
    rowvalues = []
    for cellObj in list(xlsxws.rows)[rownum]:
        cv = cellObj.value
        # process the cell values into the format required based on their values
        if cv is None:
            continue
        if 'Entry' in cv:
            continue
        if 'value' in cv:
            cv = cv.lstrip(' value=')
            cv = cv.rstrip(')')
            cv = int(cv)
            if int(cv) <= 1:
                cv = 0
        # append process cell values to row values variable
        rowvalues.append(cv)

    # write processed information into xlsx 2 file
    for columnnum, value in enumerate(rowvalues, start=1):
        xlsxws2.cell(row=(rownum + 1), column=columnnum).value = value

    # provides user with updates on how many rows have been processed
    if rownum % 100 == 0 and rownum != 0:
        print(f'Programme has processed {rownum} rows')

    # breaks loop if the lower date band specified by the user equals the current rows date value
    if lowerdate == rowvalues[0]:
        break

print(f'Programme processed a total of {rownum} rows')

# save the processed data in the output file - xlsx2
xlsx2.save(editedfilepath)
print('Processed File Saved')

print('Programme Complete')
